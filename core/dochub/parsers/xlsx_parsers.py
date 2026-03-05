import csv
import tempfile
from abc import ABC
from pathlib import Path
from typing import Dict

import openpyxl
import pandas as pd
import sqlalchemy as sa
from chardet import UniversalDetector
from loguru import logger
from openpyxl.cell import MergedCell
from pandas import MultiIndex

from dochub.parsers.base import BaseDocumentParser
from dochub.schemas import Chunk, Param, DataType, ChunkType
from utils.i18n import I18NString, Language


class BaseXlsxParser(BaseDocumentParser, ABC):
    target_content_type = [
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/wps-office.xlsx"
    ]
    target_file_ext = ["xlsx", "csv"]


class GeneralXlsxParser(BaseXlsxParser):
    name = I18NString({
        Language.ZH: "通用",
        Language.EN: "General",
    })

    params = [
        Param(name="chunk_rows", display_name="分隔行数", required=False, default=1, data_type=DataType.NUMBER)
    ]

    def __init__(self, target, **kwargs):
        super().__init__(target, **kwargs)
        self.chunk_rows = kwargs.pop("chunk_rows", 1)

    def _parse_impl(self) -> None:
        filepath = self.target.physical_path
        if self.target.doc_name.endswith("csv"):
            encoding = self.detect_encoding(self.target.physical_path)
            delimiter = self.detect_delimiter(self.target.physical_path)
            sheets = {"Sheet1": pd.read_csv(self.target.physical_path, delimiter=delimiter, encoding=encoding, encoding_errors="ignore")}
        else:
            sheets = read_excel(filepath)
        self._report_progress(20)
        sheet_items = sheets.items()
        # TODO: 优化按行解析存储逻辑
        """
        for sheet, df in sheet_items:
            for i in range(0, math.ceil(len(df) / self.chunk_rows)):
                df1 = df[i * self.chunk_rows:min(len(df), (i + 1) * self.chunk_rows)]
                if not df1.isna().all(axis=1).all():  # 过滤掉数据行为空的行
                    chunk_content = df1.to_csv(index=False, sep="|")
                    # 过滤掉空行
                    chunk_content = "\n".join(
                        [line for line in chunk_content.splitlines() if line.replace("|", "")])
                    chunk = Chunk(content=chunk_content, type=ChunkType.TABLE, metadata={"sheet": sheet})
                    self._append_content_chunks(chunk)
        self._report_progress(80)
        """
        # Add extra data chunks for saving to es index: structured_data
        db_name = Path(self.target.doc_name).stem
        sql_chunk = self.generate_sql_chunk(db_name, sheet_items)
        self._append_content_chunks(sql_chunk)

        self._report_progress(100)

    def detect_encoding(self, file_path) -> str:
        detector = UniversalDetector()
        with open(file_path, "rb") as f:
            for line in f.readlines():
                detector.feed(line)
                if detector.done:
                    break
        detector.close()
        return detector.result.get("encoding", "utf-8")

    def detect_delimiter(self, file_path, num_lines=5, encoding: str = "utf-8") -> str:
        with open(file_path, 'r', encoding=encoding, errors="ignore") as file:
            sample = ''.join([file.readline() for _ in range(num_lines)])
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            return dialect.delimiter

    def generate_sql_chunk(self, db_name, sheet_items) -> Chunk:
        # Create a temporary file-based SQLite database to ease memory pressure.
        with tempfile.NamedTemporaryFile(suffix=".db", delete=True) as temp_db:
            # Create a SQLite engine using the temporary file.
            engine = sa.create_engine(f'sqlite:///{temp_db.name}', echo=False)
            schemas = []
            for sheet_name, df in sheet_items:
                if df.empty:
                    continue
                # 每个sheet存储一张表。
                # 表名为文档名加sheet名，以防止不同文档存在同名sheet。
                table_name = f"{db_name}_{sheet_name}"
                df.to_sql(name=table_name, con=engine, index=False)

                metadata = sa.MetaData()
                table = sa.Table(table_name, metadata, autoload_with=engine)
                schemas.append(str(sa.schema.CreateTable(table)).strip())

            schema = f"\n".join(schemas)

            conn = engine.raw_connection()
            sql_dump = "\n".join(conn.iterdump())

            conn.close()
            engine.dispose()

            return Chunk(
                content=sql_dump,
                type= ChunkType.SQL,
                metadata={
                    "db_schema": schema
                }
            )


def get_merged_cells(sheet):
    merged_cells = {}
    for merged_cell_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_cell_range.bounds
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_cells[(col, row)] = sheet.cell(row=min_row, column=min_col).value
    return merged_cells


def split_excel_merged_cells(excel_path):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp_excel_file:
        # 加载Excel文件
        wb = openpyxl.load_workbook(excel_path)

        # 创建一个新的工作簿和工作表，用于存储拆分后的数据
        new_wb = openpyxl.Workbook()

        for sheet_name in wb.sheetnames:
            # 选择工作表
            work_sheet = wb[sheet_name]

            new_sheet = new_wb.create_sheet(title=sheet_name)

            # 获取合并单元格的值
            merged_cells = get_merged_cells(work_sheet)
            header_row = [cell.value for cell in work_sheet[1]]
            # 遍历工作表中的每一行（跳过表头）
            for row in work_sheet.iter_rows(min_row=work_sheet.min_row, max_row=work_sheet.max_row):
                # 遍历当前行的每个单元格
                for cell in row:
                    # 如果单元格是合并单元格的一部分，则使用存储的值
                    if (cell.column, cell.row) in merged_cells:
                        # row_values.append(merged_cells[(cell.column, cell.row)])
                        new_sheet.cell(row=cell.row, column=cell.column, value=merged_cells[(cell.column, cell.row)])
                    else:
                        # 否则，使用单元格自身的值
                        # row_values.append(cell.value)
                        new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

        # 删除默认创建的空工作表
        if new_wb.sheetnames[0] == 'Sheet':
            del new_wb[new_wb.sheetnames[0]]

        new_wb.save(tmp_excel_file)
        sheet_items = pd.read_excel(tmp_excel_file, sheet_name=None).items()
        return sheet_items


def read_excel(excel_path) -> Dict[str, pd.DataFrame]:
    # 加载Excel文件
    logger.info(f"Reading excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    dfs = {}
    for sheet_name in wb.sheetnames:
        # 选择工作表
        work_sheet = wb[sheet_name]
        logger.info(f"Reading sheet: {sheet_name}")

        if work_sheet.max_row == 1 and work_sheet.max_column == 1 and work_sheet['A1'].value is None:
            logger.info(f"Skipped empty sheet: {sheet_name}")
            continue

        # 展开合并单元格
        # 获取合并单元格的值
        if merged_cells := get_merged_cells(work_sheet):
            logger.info("Flattening merged cells...")
            work_sheet.title = f"Deprecated_{sheet_name}"
            new_sheet = wb.create_sheet(title=sheet_name)

            # 遍历工作表中的每一行
            for row in work_sheet.iter_rows():
                # 遍历当前行的每个单元格
                for cell in row:
                    # 如果单元格是合并单元格的一部分，则使用存储的值
                    if (cell.column, cell.row) in merged_cells:
                        new_sheet.cell(row=cell.row, column=cell.column, value=merged_cells[(cell.column, cell.row)])
                    else:
                        # 否则，使用单元格自身的值
                        new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

            del wb[work_sheet.title]
            work_sheet = new_sheet

        header = detect_header(work_sheet)
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header)

        # 丢弃数值全部为NaN的空行/列
        df = df.dropna(axis=0, how='all')
        df = df.dropna(axis=1, how='all')

        # 展开多行表头
        if isinstance(df.columns, MultiIndex):
            flat_index = df.columns.to_flat_index()
            cleaned_flat_index = []
            for i in flat_index:
                # 表头去重
                deduplicates = []
                for name in i:
                    if name not in deduplicates:
                        deduplicates.append(name)

                # 去除无意义的空表头
                meaningful = []
                for name in deduplicates:
                    if isinstance(name, str) and name.startswith('Unnamed'):
                        continue
                    meaningful.append(name)

                # 用'_'拼接多维表头
                cleaned_flat_index.append('_'.join([str(x) for x in meaningful]))
            df.columns = cleaned_flat_index

        if pd.api.types.is_string_dtype(df.columns):
            df.columns = df.columns.str.strip()

        dfs[sheet_name] = df

    return dfs


def detect_header(work_sheet, max_data_cell_ratio=0.1):
    """
    Detects the actual header of an Excel sheet.
    """
    header = []

    cell_stats = []
    # 遍历每一行
    for i, row in enumerate(work_sheet.iter_rows()):
        total_cell_count = len(row)
        data_cell_count = 0
        str_cell_count = 0
        str_groups = []
        last_str_value = None
        not_na_values = set()

        # 遍历当前行的每个单元格
        for cell in row:
            if cell.value is not None:
                not_na_values.add(cell.value)
            if is_data_cell(cell):
                data_cell_count += 1
            if cell.data_type == "s":
                str_cell_count += 1
                if cell.value != last_str_value:
                    if last_str_value is not None:
                        str_groups.append(last_str_value)
                    last_str_value = cell.value
        if str_groups and last_str_value != str_groups[-1]:
            str_groups.append(last_str_value)

        if data_cell_count == str_cell_count == 0:
            logger.info(f"Row {i} skipped due to empty.")
            continue

        if len(not_na_values) == 1:
            logger.info(f"Row {i} skipped due to single value: {not_na_values.pop()}")
            continue

        data_cell_ratio = data_cell_count / total_cell_count
        str_cell_ratio = str_cell_count / total_cell_count

        # 如果一行中的数据单元格数量小于单元格总数的一定比例（默认10%）时，认为该行为表头行。
        if data_cell_ratio <= max_data_cell_ratio:
            # 如果一行内字符型单元格占比小于上一行的字符型单元格占比，或该行的字符型数据不如上一行丰富，则认为该行不会是接续上一行的低维表头，应当立即停止寻找表头。
            if header and cell_stats and (cell_stats[-1][1] >= str_cell_ratio and len(cell_stats[-1][2]) >= len(str_groups)):
                logger.info(f"Row {i} could not be a subsequent header row. ({data_cell_ratio * 100:.2f}% data cells, {str_cell_ratio * 100:.2f}% str cells, {len(str_groups)} str groups)")
                break

            logger.info(f"Row {i} is probably a{' subsequent' if header else ''} header row. ({data_cell_ratio * 100:.2f}% data cells, {str_cell_ratio * 100:.2f}% str cells, {len(str_groups)} str groups)")
            header.append(i)
        else:
            logger.info(f"Row {i} is probably a data row. ({data_cell_ratio * 100:.2f}% data cells, {str_cell_ratio * 100:.2f}% str cells)")

            # 遇到数据行时，如果此时还没有检测到表头，并且上一行的数据型单元格占比没有超过阈值，则将上一行视为表头行，否则直接停止
            if not header and cell_stats and cell_stats[-1][0] <= max_data_cell_ratio:
                logger.info(f"Row {i - 1} is regarded as a header row due to empty headers. ({cell_stats[-1][0] * 100:.2f}% data cells, {cell_stats[-1][1] * 100:.2f}% str cells)")
                header.append(i - 1)

            break

        cell_stats.append((data_cell_ratio, str_cell_ratio, str_groups))

    logger.info(f"Header detected: {header}")
    return header or None


def is_data_cell(cell):
    if isinstance(cell, MergedCell):
        return False
    if cell.value is None:
        return False
    return cell.is_date or cell.data_type in ['n', 'b', 'e', 'f']
